import io, os, csv, json, time
from datetime import datetime, time as dtime, timedelta
from typing import List, Dict, Any, Optional

import httpx
import uvicorn
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook

# Импорт нового модуля оптимизации маршрута
from route_optimizer_2gis import RouteOptimizer2GIS

# ───────── CONFIG
YANDEX_API_KEY = os.getenv("YANDEX_API_KEY", "58c38b72-57f7-4946-bc13-a256d341281a").strip()
DGIS_API_KEY = os.getenv("DGIS_API_KEY", "09e6cea9-9540-4665-934e-1864124b7304").strip()
GEOCODE_CACHE_PATH = os.getenv("GEOCODE_CACHE_PATH", "./geocode_cache.json")
GEOCODE_SLEEP_SEC = float(os.getenv("GEOCODE_SLEEP_SEC", "0.15"))
HTTP_TIMEOUT = float(os.getenv("HTTP_TIMEOUT", "10"))

app = FastAPI(title="Planner API (matrix + chain modes)", version="0.5.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ───────── Utils
def parse_time(s: Any) -> Optional[dtime]:
    if s is None: return None
    if isinstance(s, dtime): return s
    s = str(s).strip().replace('.', ':')
    if not s: return None
    if s.isdigit() and len(s) in (3,4):
        s = s.zfill(4); s = s[:2] + ':' + s[2:]
    try:
        hh, mm = s.split(':'); return dtime(int(hh), int(mm))
    except Exception:
        return None

def today_at(t: Optional[dtime]) -> Optional[datetime]:
    if t is None: return None
    now = datetime.now()
    return datetime(now.year, now.month, now.day, t.hour, t.minute, 0)

def parse_iso_any(s: Optional[str]) -> Optional[datetime]:
    if not s: return None
    s = str(s)
    try:
        if s.endswith('Z'):
            dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
        else:
            dt = datetime.fromisoformat(s)
    except Exception:
        return None
    return dt if dt.tzinfo is None else dt.astimezone().replace(tzinfo=None)

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6371000.0
    from math import radians, sin, cos, atan2, sqrt
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dlambda/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R * c

# ───────── IO (CSV/XLSX)
def read_csv(content: bytes) -> List[Dict[str, Any]]:
    import csv as _csv, io as _io
    text = content.decode('utf-8-sig', errors='replace')
    rdr = _csv.reader(_io.StringIO(text))
    rows = list(rdr)
    if not rows: return []
    header = [h.strip() for h in rows[0]]
    out = []
    for i, row in enumerate(rows[1:], start=2):
        rec = {header[j].strip().lower(): (row[j] if j < len(row) else '') for j in range(len(header))}
        rec['_row'] = i
        out.append(rec)
    return out

def read_xlsx(content: bytes) -> List[Dict[str, Any]]:
    bio = io.BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for i, r in enumerate(rows[1:], start=2):
        rec = {}
        for j, h in enumerate(header):
            key = h.strip().lower()
            rec[key] = r[j] if j < len(r) else None
        rec['_row'] = i
        out.append(rec)
    return out

def load_records(file: UploadFile) -> List[Dict[str, Any]]:
    content = file.file.read()
    name = (file.filename or '').lower()
    if name.endswith('.csv') or file.content_type in ('text/csv','application/csv'):
        return read_csv(content)
    else:
        return read_xlsx(content)

# ───────── Геокодер (кэш)
def _load_cache() -> Dict[str, Any]:
    if os.path.exists(GEOCODE_CACHE_PATH):
        try:
            with open(GEOCODE_CACHE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_cache(cache: Dict[str, Any]) -> None:
    try:
        with open(GEOCODE_CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

_geocode_cache = _load_cache()

def geocode_address(address: str) -> Optional[Dict[str, float]]:
    if not address: return None
    if address in _geocode_cache:  # cache hit
        return _geocode_cache[address]
    if not YANDEX_API_KEY:
        return None

    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {"apikey": YANDEX_API_KEY, "format": "json", "geocode": address}
    try:
        with httpx.Client(timeout=HTTP_TIMEOUT) as client:
            r = client.get(url, params=params)
            r.raise_for_status()
            data = r.json()
            items = data.get("response", {}).get("GeoObjectCollection", {}).get("featureMember", [])
            if not items: return None
            pos = items[0]["GeoObject"]["Point"]["pos"]  # "lon lat"
            lon, lat = map(float, pos.split())
            res = {"lat": lat, "lon": lon}
            _geocode_cache[address] = res
            _save_cache(_geocode_cache)
            time.sleep(GEOCODE_SLEEP_SEC)
            return res
    except Exception:
        return None

# ───────── Канонизация строк файла
def canonicalize(records: List[Dict[str, Any]], default_service_min: int = 20, do_geocode: bool = True) -> List[Dict[str, Any]]:
    def v(rec: Dict[str, Any], *alts):
        for a in alts:
            if a in rec and rec[a] not in (None, ''):
                return rec[a]
        return None

    out = []
    for i, rec in enumerate(records):
        rid = v(rec, 'номер объекта', 'id', 'номер')
        try: rid = int(str(rid).split('.')[0]) if rid is not None else (i + 1)
        except: rid = i + 1

        addr = str(v(rec, 'адрес объекта', 'адрес', 'address') or '').strip()

        def to_float(x):
            try: return float(str(x).replace(',', '.'))
            except: return None

        lat = to_float(v(rec, 'географическая широта', 'широта', 'lat', 'latitude'))
        lon = to_float(v(rec, 'географическая долгота', 'долгота', 'lon', 'lng', 'longitude'))

        if (lat is None or lon is None) and addr and do_geocode:
            geo = geocode_address(addr)
            if geo: lat, lon = geo['lat'], geo['lon']

        ws = parse_time(v(rec, 'время начала рабочего дня', 'окно начало', 'окно_начало', 'window_start'))
        we = parse_time(v(rec, 'время окончания рабочего дня', 'окно конец', 'окно_конец', 'window_end'))
        lvl = str(v(rec, 'уровень клиента', 'приоритет') or '').strip().lower()
        service = default_service_min if lvl != 'vip' else max(default_service_min, 25)

        row = {
            "id": rid,
            "client": f"Клиент {rid}",
            "address": addr,
            "lat": lat, "lon": lon,
            "window_start": today_at(ws).isoformat() if ws else None,
            "window_end": today_at(we).isoformat() if we else None,
            "priority": lvl,
            "service_min": service
        }
        out.append(row)
    return [r for r in out if r["lat"] is not None and r["lon"] is not None]

# ───────── Модели запросов
class Point(BaseModel):
    id: Optional[int] = None
    lat: float
    lon: float
    address: Optional[str] = None
    window_start: Optional[str] = None
    window_end: Optional[str] = None
    service_min: Optional[int] = 20

class MatrixOptimizeRequest(BaseModel):
    points: List[Point]
    duration_matrix_sec: List[List[float]]
    start_time: Optional[str] = "09:00"
    start_index: int = 0

class ChainOptimizeRequest(BaseModel):
    """Точки уже в порядке обхода; legs_sec[i] — длительность от точки i-1 к i (для i=0 = 0)."""
    points: List[Point]
    legs_sec: List[float]
    start_time: Optional[str] = "09:00"

# ───────── Расчёт по цепочке
def build_route_with_chain(points: List[Point], legs_sec: List[float], start_time: Optional[str]) -> Dict[str, Any]:
    n = len(points)
    if n == 0:
        return {"summary":{"drive_min":0,"service_min":0,"wait_min":0,"total_elapsed_min":0,
                           "total_time_min":0,"visits":0,"on_time":0,"late":0,"late_penalty":0},
                "route":[]}
    if len(legs_sec) != n:
        raise HTTPException(status_code=400, detail="Длина legs_sec должна равняться числу точек (первый элемент = 0).")

    st = parse_time(start_time) if start_time else dtime(9, 0)
    cur = today_at(st)

    drive = wait = service = 0.0
    on_time = late = penalty = 0
    out = []

    for i, p in enumerate(points):
        # проезд между i-1 -> i
        tt_sec = float(legs_sec[i] or 0.0)
        cur = cur + timedelta(seconds=tt_sec)
        drive += tt_sec / 60.0

        ws = parse_iso_any(p.window_start)
        we = parse_iso_any(p.window_end)

        lm = 0
        if ws and cur < ws:
            w = (ws - cur).total_seconds() / 60.0
            wait += w
            cur = ws
        elif we and cur > we:
            lm = int(round((cur - we).total_seconds() / 60.0))

        if lm <= 0: on_time += 1
        else:
            late += 1
            penalty += max(0, lm - 5)

        svc = int(p.service_min or 20)
        service += svc

        out.append({
            "id": p.id if p.id is not None else (i + 1),
            "client": f"Клиент {p.id}" if p.id is not None else f"Клиент {i+1}",
            "address": p.address or "",
            "lat": p.lat, "lon": p.lon,
            "window_start": p.window_start, "window_end": p.window_end,
            "service_min": svc,
            "eta": cur.isoformat(),
            "late_minutes": lm
        })

        cur = cur + timedelta(minutes=svc)

    summary = {
        "drive_min": int(round(drive)),
        "service_min": int(round(service)),
        "wait_min": int(round(wait)),
        "total_elapsed_min": int(round(drive + service + wait)),
        "total_time_min": int(round(drive)),
        "visits": len(out),
        "on_time": on_time,
        "late": late,
        "late_penalty": int(penalty)
    }
    return {"summary": summary, "route": out}

# ───────── API
@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/api/validate_dgis_key")
async def validate_dgis_key(dgis_api_key: str = Query(..., description="Ключ API 2ГИС для проверки")):
    """
    Проверяет валидность ключа API 2ГИС.
    """
    try:
        # Валидация формата ключа
        if not dgis_api_key or not dgis_api_key.strip():
            return {"valid": False, "error": "Ключ API не может быть пустым"}
        
        import re
        key_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
        if not re.match(key_pattern, dgis_api_key.strip(), re.IGNORECASE):
            return {"valid": False, "error": "Неверный формат ключа API. Ожидается UUID формат"}
        
        # Проверяем ключ через создание клиента и тестовый запрос
        from dgis_matrix_client import DGISMatrixClient, DGISMatrixError
        
        try:
            client = DGISMatrixClient(dgis_api_key.strip())
            client.warmup()  # Тестовый запрос
            return {"valid": True, "message": "Ключ API валиден"}
        except DGISMatrixError as e:
            return {"valid": False, "error": f"Ошибка API 2ГИС: {str(e)}"}
        except Exception as e:
            return {"valid": False, "error": f"Неожиданная ошибка: {str(e)}"}
            
    except Exception as e:
        return {"valid": False, "error": f"Ошибка валидации: {str(e)}"}

@app.post("/api/optimize")
async def optimize(
    file: UploadFile = File(...),
    start_time: Optional[str] = Query(default="09:00"),
    speed_kmh: float = Query(default=25.0, ge=1.0, le=120.0),
    default_service_min: int = Query(default=20, ge=0, le=240)
):
    try:
        records = load_records(file)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {e}")

    rows = canonicalize(records, default_service_min=default_service_min, do_geocode=True)
    if not rows:
        raise HTTPException(status_code=400, detail="Не найдено ни одной строки с координатами (и геокод не помог).")

    # простая псевдо-матрица — обратная совместимость
    pts = [{"lat": r["lat"], "lon": r["lon"]} for r in rows]
    n = len(pts)
    mat_sec = [[0.0]*n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i == j: continue
            d = haversine_m(pts[i]["lat"], pts[i]["lon"], pts[j]["lat"], pts[j]["lon"])
            t_min = (d/1000.0) / max(speed_kmh,1e-3) * 60.0
            mat_sec[i][j] = t_min * 60.0

    # построим тривиальную цепочку по порядку
    legs = [0.0] + [mat_sec[i-1][i] for i in range(1, n)]
    points = [Point(**{
        "id": r["id"], "lat": r["lat"], "lon": r["lon"], "address": r["address"],
        "window_start": r["window_start"], "window_end": r["window_end"],
        "service_min": r["service_min"]
    }) for r in rows]
    return build_route_with_chain(points, legs, start_time)

@app.post("/api/optimize_with_matrix")
async def optimize_with_matrix(req: MatrixOptimizeRequest):
    # Оставлен на случай, если понадобится полный TSP по матрице
    points = req.points
    n = len(points)
    if n == 0:
        return {"summary":{"drive_min":0,"service_min":0,"wait_min":0,"total_elapsed_min":0,
                           "total_time_min":0,"visits":0,"on_time":0,"late":0,"late_penalty":0},
                "route":[]}
    if len(req.duration_matrix_sec) != n or any(len(r)!=n for r in req.duration_matrix_sec):
        raise HTTPException(status_code=400, detail="Матрица должна быть NxN.")
    # Превратим матрицу в «цепочку» по порядку NN (простой, локальный выбор)
    remaining = set(range(n))
    cur = 0
    order = [cur]; remaining.remove(cur)
    while remaining:
        nxt = min(remaining, key=lambda j: float(req.duration_matrix_sec[cur][j] or 1e18))
        order.append(nxt); remaining.remove(nxt); cur = nxt
    legs = [0.0]
    for i in range(1, n):
        legs.append(float(req.duration_matrix_sec[order[i-1]][order[i]] or 0.0))
    ordered_points = [points[i] for i in order]
    return build_route_with_chain(ordered_points, legs, req.start_time)

@app.post("/api/optimize_with_chain")
async def optimize_with_chain(req: ChainOptimizeRequest):
    if not req.points or not req.legs_sec:
        raise HTTPException(status_code=400, detail="Нужны points и legs_sec.")
    return build_route_with_chain(req.points, req.legs_sec, req.start_time)

@app.post("/api/optimize_2gis")
async def optimize_2gis(
    file: UploadFile = File(...),
    dgis_api_key: str = Query(..., description="Ключ API 2ГИС"),
    work_start: str = Query(default="09:00", description="Время начала работы (HH:MM)"),
    work_end: str = Query(default="18:00", description="Время окончания работы (HH:MM)"),
    meeting_minutes: int = Query(default=30, ge=5, le=240, description="Длительность визита в минутах")
):
    """
    Новый endpoint для оптимизации маршрута с использованием 2ГИС API.
    Использует реальные данные о пробках и более точное геокодирование.
    """
    try:
        # Сохраняем загруженный файл во временный файл
        import tempfile
        with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.csv') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        try:
            # Валидация ключа API
            if not dgis_api_key or not dgis_api_key.strip():
                raise HTTPException(status_code=400, detail="Ключ API 2ГИС не может быть пустым")
            
            # Базовая валидация формата ключа (UUID формат)
            import re
            key_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
            if not re.match(key_pattern, dgis_api_key.strip(), re.IGNORECASE):
                raise HTTPException(status_code=400, detail="Неверный формат ключа API 2ГИС. Ожидается UUID формат")
            
            # Создаем оптимизатор с переданным ключом 2ГИС
            optimizer = RouteOptimizer2GIS(dgis_api_key.strip())
            
            # Запускаем оптимизацию
            result = optimizer.optimize_from_csv(
                csv_path=tmp_file_path,
                work_start_str=work_start,
                work_end_str=work_end,
                meeting_minutes=meeting_minutes
            )
            
            # Преобразуем результат в формат, совместимый с существующим API
            converted_result = convert_2gis_result_to_api_format(result)
            
            return converted_result
            
        finally:
            # Удаляем временный файл
            try:
                os.unlink(tmp_file_path)
            except:
                pass
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка оптимизации маршрута: {str(e)}")

def convert_2gis_result_to_api_format(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Преобразует результат оптимизации 2ГИС в формат, совместимый с существующим API.
    """
    schedule = result.get("schedule", [])
    
    # Преобразуем расписание в формат API
    route = []
    for item in schedule:
        route_item = {
            "id": item.get("id"),
            "client": f"Клиент {item.get('id')}",
            "address": item.get("address", ""),
            "lat": None,  # Координаты не возвращаются в текущем формате
            "lon": None,
            "window_start": None,  # Временные окна не возвращаются
            "window_end": None,
            "service_min": result.get("meeting_minutes", 30),
            "eta": item.get("arrive"),
            "late_minutes": 0  # Пока не вычисляется
        }
        route.append(route_item)
    
    # Создаем summary в формате API
    stats = result.get("stats", {})
    summary = {
        "drive_min": int(stats.get("total_drive_min", 0)),
        "service_min": len(schedule) * result.get("meeting_minutes", 30),
        "wait_min": 0,  # Пока не вычисляется
        "total_elapsed_min": int(stats.get("total_drive_min", 0)) + len(schedule) * result.get("meeting_minutes", 30),
        "total_time_min": int(stats.get("total_drive_min", 0)),
        "visits": len(schedule),
        "on_time": len(schedule),  # Пока считаем все вовремя
        "late": 0,
        "late_penalty": 0
    }
    
    return {
        "summary": summary,
        "route": route,
        "external_provider": result.get("external_provider", "2GIS"),
        "uses_external_api": result.get("uses_external_api", True),
        "total_distance_km": stats.get("total_distance_km", 0),
        "work_window": result.get("work_window", {}),
        "visited_count": stats.get("visited_count", 0),
        "dropped_count": stats.get("dropped_count", 0)
    }

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
